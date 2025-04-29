import time

from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By

url = "https://www.saucedemo.com/"
file = 'testdata.xlsx'
wb = load_workbook(file)     # load_workbook function from openpyxl
ws = wb.active

# Reading Username and Password from file
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    username = row[0]
    password = row[1]

    browser = webdriver.Firefox()
    browser.maximize_window()
    browser.get(url)

    time.sleep(2)

    browser.find_element(By.CSS_SELECTOR, "#user-name").send_keys(username)
    browser.find_element(By.XPATH, "//input[@id='password']").send_keys(password)
    browser.find_element(By.CSS_SELECTOR, "#login-button").click()
    time.sleep(2)

    browser.find_element(By.CSS_SELECTOR, "#react-burger-menu-btn").click()
    browser.find_element(By.XPATH, "//a[@id='logout_sidebar_link']").click()

    time.sleep(3)

    browser.quit()








